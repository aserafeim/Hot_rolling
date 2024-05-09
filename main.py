from deformation import dislocation_density_calc, strain_calc
import matplotlib.pyplot as plt
import openpyxl
import sys
import static_recrystallisation
import inputdata
import material_constants as mat_const
import dynamic_recrystallisation as drx
import math
import time
from plotting import plot_func


print('sys.path=', sys.path)
location = r'C:\Users\aserafeim\Desktop\python_HR\Test_srx.xlsx'
wb = openpyxl.load_workbook(location)
sheet_curr = wb['Static_RX']
max_col = sheet_curr.max_column
max_row = sheet_curr.max_row
main_dict = {}

'''reading excel file'''
for i in range(3, max_row+1):
    temp_dict = {}
    temp_dict['time_increment'] = sheet_curr.cell(i, 2).value
    temp_dict['initial_strain'] = sheet_curr.cell(i, 3).value
    temp_dict['final_strain'] = sheet_curr.cell(i, 4).value
    temp_dict['initial_strain_rate'] = sheet_curr.cell(i, 5).value
    temp_dict['final_strain_rate'] = sheet_curr.cell(i, 6).value
    temp_dict['initial_temp'] = sheet_curr.cell(i, 7).value+273.15
    temp_dict['final_temp'] = sheet_curr.cell(i, 8).value+273.15
    temp_dict['rho_0'] = sheet_curr.cell(i, 9).value if (isinstance(sheet_curr.cell(i, 9).value, int) or
                                                         isinstance(sheet_curr.cell(i, 9).value, float)) else 0.0
    temp_dict['pass_interval'] = sheet_curr.cell(i, 10).value
    main_dict[str(i-2)] = temp_dict
for key, items in main_dict.items():
    print(key, ':', items, '\n')

print('"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""')
#print(max_row)


counter = 0
start = time.time()
'''running of loops for passes'''
for i in range(1, max_row-1):
    
    ''' initialisation of variables'''
    total_time = []
    variable_pass = main_dict[str(i)]
    print(i)
    if i == 1:
        dislocation_density_initial = variable_pass['rho_0'] 
    else:
        if main_dict[str(i-1)]['initial_strain_rate'] == 0:
            dislocation_density_initial = main_dict[str(i-1)]['rho_m_list'][-1]
            #dislocation_density_initial = 1e+12
        else:
            dislocation_density_initial = main_dict[str(i-1)]['rho_list'][-1]
            #dislocation_density_initial = 1e+14

    time_step = variable_pass['time_increment']
    strain_initial = variable_pass['initial_strain']
    pass_interval = variable_pass['pass_interval']
    time_curr = time_step
    rho_curr = dislocation_density_initial
    strain_rate_initial = variable_pass['initial_strain_rate']
    strain_rate_temp = strain_rate_initial
    temp_init = variable_pass['initial_temp']
    temp_final = variable_pass['final_temp']

    'Tracking the temperature change'
    temp_rate = (temp_final - temp_init)/pass_interval
    #print('temp rate:', temp_rate)
    ''' last_time used for real time calculation and global time calculation'''
    if i > 1:
        last_time = main_dict[str(i-1)]['global_time'][-1]
        #print('last time:', last_time)
    temp_curr = temp_init

    time_list = []
    strain_list, temperature_list = [], []
    number_timesteps = pass_interval / time_step
    rho_cr_list = []
    
    '''checks if strain rate is 0 or not, i.e deformation or RX'''
    if strain_rate_initial > 0:
        temp_strain = strain_initial
        rho_list, drx_rxc_list, drx_xcurr_list, drx_rholist, drx_rrx_list, drx_rgcurr_list, drx_nrx_list, drx_stress_list = [],[],[],[],[],[],[],[]
        time_arr, dn_rx_dt = [], []

        while time_curr <= pass_interval:
            temp_strain = strain_calc(temp_strain, time_step, strain_rate_initial)
            #rho_temp_var = 0
            #acc_strain += temp_strain
            'calculation of rho critical'
            n_z = 0.1 if strain_rate_temp > 1 else 1
            D_t = inputdata.grain_size_init * math.exp(-2*temp_strain/math.sqrt(3))
            gb_energy = mat_const.calc_austenite_gb_energy(temp_curr)
            M_g = mat_const.calc_gb_mobility2_drx(temp_curr)
            part1 = (16*mat_const.calc_austenite_gb_energy(temp_curr) * math.pow(strain_rate_temp,n_z) * inputdata.M) / (3*mat_const.calc_burgers_vector(temp_curr)*mat_const.calc_gb_mobility2_drx(temp_curr)*math.pow(mat_const.calc_dislocation_line_energy(temp_curr),2))
            part2 = (inputdata.fitting_params['C_1'] / D_t) + (inputdata.fitting_params['C_2'] * math.sqrt(inputdata.rho_const))
            rho_critical = math.pow(part1*part2, 1/3)
            
            #print('rho_critical:', rho_critical)
            
            'Call to the deformation module'
            temp_dislocation_density = dislocation_density_calc(rho_curr, time_step, temp_curr, strain_rate_initial, D_t)
            rho_temp_var = temp_dislocation_density
            
            'Adding values to the lists of temp, strain and dislocation_density'
            time_list.append(time_curr)
            rho_list.append(temp_dislocation_density)
            temperature_list.append(temp_curr)
            strain_list.append(temp_strain)

            if i == 1:
                total_time.append(time_curr)
            else:
                total_time.append(last_time+time_curr)
                
            'checking for dynamic recrystallisation'
            if temp_dislocation_density > rho_critical:
                #print('inside DRX')
                counter += 1
                #print('inside dynamic rx function')
                if counter == 1:
                    nrx_drx = 0
                    x_prev = 0
                    x_c_prev=0
                    r_rx_prev, r_g_prev = 0,0
                    #print('out DRX')
                dyn_rx_dict = drx.dynamic_rx(rho_critical, temp_dislocation_density, nrx_drx, r_rx_prev, r_g_prev, x_prev, x_c_prev, time_step, temp_curr, strain_rate_temp,temp_strain)
                
                'appending values to the lists'
                drx_rxc_list.append(dyn_rx_dict['r_xc'])
                drx_xcurr_list.append(dyn_rx_dict['x_curr'])
                drx_rholist.append(dyn_rx_dict['rho_m'])
                drx_rrx_list.append(dyn_rx_dict['r_rx'])
                drx_rgcurr_list.append(dyn_rx_dict['r_g_curr'])
                drx_nrx_list.append(dyn_rx_dict['n_rx'])
                time_arr.append(time_curr)
                rho_temp_var = dyn_rx_dict['rho_m']
                rho_cr_list.append(rho_critical)
                dn_rx_dt.append(dyn_rx_dict['nucleation_rate'])
                
                
                'updating values after every iteration'
                nrx_drx = dyn_rx_dict['n_rx']
                r_rx_prev = dyn_rx_dict['r_rx']
                r_g_prev = dyn_rx_dict['r_g_curr']
                x_prev = dyn_rx_dict['x_curr']
                x_c_prev=dyn_rx_dict['x_c_curr']
                
                if dyn_rx_dict['x_curr']> 0.99:
                    plot_func(time_arr, drx_xcurr_list, 'Time', 'drx_xcurr', True, False, variable_pass)
                    break
                   
            'calculation of stress'
            stress_drx = mat_const.calc_stress(temp_curr, strain_rate_temp, rho_temp_var)
            drx_stress_list.append(stress_drx)
            
            'Updating variables for next iteration for the whole sub-routine'
            prev_stress = stress_drx
            time_curr += time_step
            rho_curr = temp_dislocation_density
            temp_curr = temp_curr + temp_rate * time_step
            
            '''temp_curr = temp_init - (temp_final-temp_init)/number_timesteps'''
            
        

        variable_pass['time_list'] = time_list
        variable_pass['rho_list'] = rho_list
        variable_pass['strain_list'] = strain_list
        variable_pass['temperature_list'] = temperature_list
        variable_pass['global_time'] = total_time
        variable_pass['stress_list'] = drx_stress_list
        variable_pass['rho_cr_list'] = rho_cr_list
        variable_pass['drx_xcurr_list'] = drx_xcurr_list
        print(drx_xcurr_list[-1])

    else:
        #print('dislocation_density_initial inside srx:', dislocation_density_initial)
        radius_list, rho_m_list, no_rx_grains_list, rx_fraction_list, rrx_prev_crit = [], [], [], [], []
        rho_def = dislocation_density_initial
        dn_rx_list, new_rx_size, driving_force, mean_rx_grainsize, nuclei_radius = [], [], [], [],[]
#         = []
        print('inside static rx')
        'Static recrystallisation module runs'
        r_rx_prev = 2.5e-4
        while time_curr <= pass_interval:
            d0 = inputdata.grain_size_init
           #p_grain_size = 1e-6
            'special conditions for the first iteration'
            if time_curr == time_step:
                no_grains_rx = 1.0
                radius_grains = 1e-6 
                rho_m = dislocation_density_initial
                
            temp_srx_dict = static_recrystallisation.static_rx(rho_m, rho_def, no_grains_rx, radius_grains, d0, temp_curr, time_step, r_rx_prev)
            
            
            'adding values to the list, later used for plotting the graphs'
            time_list.append(time_curr)
            temperature_list.append(temp_curr)
            radius_list.append(temp_srx_dict['mean_grain_size'])
            rho_m_list.append(temp_srx_dict['mean_rho_new'])
            no_rx_grains_list.append(temp_srx_dict['N_rx'])
            rx_fraction_list.append(temp_srx_dict['recrystallised_fraction'])
            dn_rx_list.append(temp_srx_dict['newly_nucleated_grains'])
            new_rx_size.append(temp_srx_dict['new_rxed_grain_size'])
            driving_force.append(temp_srx_dict['driving_force'])
            mean_rx_grainsize.append(temp_srx_dict['mean_rxed_grain_size'])
            nuclei_radius.append(temp_srx_dict['critical_nuclei_radius'])
            rrx_prev_crit.append(temp_srx_dict['r_rx_prev-critical'])
            'tracking real time'
            if i == 1:
                total_time.append(time_curr)
            else:
                total_time.append(last_time+time_curr)
            
            'updating values after every iteration'
            time_curr += time_step
            r_rx_prev = temp_srx_dict['mean_rxed_grain_size']
            radius_grains = temp_srx_dict['new_rxed_grain_size']
            no_grains_rx = temp_srx_dict['N_rx']
            rho_m = temp_srx_dict['mean_rho_new']
            temp_curr = temp_curr + temp_rate * time_step

        variable_pass['time_list'] = time_list
        variable_pass['temperature_list']=temperature_list
        variable_pass['radius_list']=radius_list # mean grain size
        variable_pass['rho_m_list']=rho_m_list
        variable_pass['no_rx_grains_list']= no_rx_grains_list
        variable_pass['rx_fraction_list'] = rx_fraction_list
        variable_pass['global_time'] = total_time
        variable_pass['dn_rx_list'] = dn_rx_list
        #variable_pass['new_rxed_grain_size'] = new_rxed_grain_size
        variable_pass['new_rxed_grain_size'] = new_rx_size #Rg_i
        variable_pass['driving_force'] = driving_force
        variable_pass['mean_rx_grainsize'] = mean_rx_grainsize #r_rx
        variable_pass['nuclei_radius'] = nuclei_radius #criticial nuclei Gmama/ G
        variable_pass['rrx_prev_crit'] = rrx_prev_crit
        print(rx_fraction_list[-1])

total_strain_list = []
total_rho_list = []
total_temp_list = []
real_time = []
rx_fraction = []
time_each = []
number = []
grain_size = []
dnrx, r_rx = [], []
rg, nuclei, driving_force = [], [], []
rrx_prev_crit_arr, rho_c = [], []
d_timelist, dislocation_list = [],[]
stress_list, strain_list, drx_xcurr_list = [],[], []

for key, val in main_dict.items():
    #total_strain_list += val['strain_list']
    #total_rho_list += val['rho_list']
    total_temp_list += val['temperature_list']
    real_time += val['global_time']
    if 'rx_fraction_list' in val:
        time_each.extend(val['time_list'])
        number.extend(val['no_rx_grains_list'])
        total_rho_list.extend(val['rho_m_list'])
        rx_fraction.extend(val['rx_fraction_list'])
        grain_size.extend(val['radius_list'])
        dnrx.extend(val['dn_rx_list'])
        rg.extend(val['new_rxed_grain_size'])
        nuclei.extend(val['nuclei_radius'])
        driving_force.extend(val['driving_force'])
        r_rx.extend(val['mean_rx_grainsize'])
        rrx_prev_crit_arr.extend(val['rrx_prev_crit'])
    else:
        #print('deformation module rho end: ', val['rho_list'][-1] )
        rho_c.extend(val['rho_cr_list'])
        d_timelist.extend(val['time_list'])
        dislocation_list.extend(val['rho_list'])
        stress_list.extend(val['stress_list'])
        strain_list.extend(val['strain_list'])
        drx_xcurr_list.extend(val['drx_xcurr_list'])
        print('stress', val['stress_list'][-1]/1E+06)
#print(nuclei)
#plt.xlim(0,100)
#plt.ylim(0,1)
#print(max(stress_list))
# =============================================================================
# plt.plot(d_timelist, drx_xcurr_list, 'r')
# #plt.plot(d_timelist, dislocation_list, 'b')
# plt.xscale('log')
# plt.grid(True,which="both", linestyle='--')
# plt.xlabel('time')
# plt.ylabel('rho')
# plt.title('rho vs time')
# 
# plt.show()
# =============================================================================
#plot_func(d_timelist, dislocation_list, 'Time', 'rho_list', True, False, main_dict)
end = time.time()
print('Time elapsed', end - start)
